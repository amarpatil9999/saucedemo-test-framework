import openpyxl

workbook = openpyxl.load_workbook("test_data/data.xlsx")
sheet = workbook['Sheet1']

# To get row size in sheet
rowSize = sheet.max_row
print("Row Size =",rowSize)

# To get column size in sheet
colSize = sheet.max_column
print("Column Size =",colSize)

# To get data from sheet1
# Approach 1
s1 = sheet['A1'].value
print("Value =",s1)

# Approach 2
s2 = sheet.cell(row=1,column=1).value
print("Value =", s2)

print(sheet.cell(row=1,column=1).value)