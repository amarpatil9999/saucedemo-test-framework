import openpyxl

class ReadTestData:
    @staticmethod
    def getTestData(rowIndex,colIndex):
        workbook = openpyxl.load_workbook("test_data/data.xlsx")
        sheet = workbook['Sheet2']
        s2 = sheet.cell(row=rowIndex,column=colIndex).value
        return s2

    @staticmethod
    def getTestDataInt(rowIndex,colIndex):
        workbook = openpyxl.load_workbook("test_data/data.xlsx")
        sheet = workbook['Sheet2']
        s2 = sheet.cell(row=rowIndex, column=colIndex).value
        s2 = int(s2)    #convert String to Int
        return s2
